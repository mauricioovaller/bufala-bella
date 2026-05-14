import pandas as pd
import numpy as np
import math
from datetime import date

# ─────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────
FECHA_MINIMA      = date(2025, 1, 1)    # 01/01/2025 (mínimo Fecha Inicial)
FECHA_MAX_RETIRO  = date(2025, 12, 30)  # 30/12/2025 (cuando Fecha Retiro está vacía)
SALARIO_DIARIO    = 1_423_500 / 30
AUX_DIARIO        = 200_000  / 30
TOPE_PREST        = 0.2183              # tope prestaciones sobre (Salario+Aux)

# Documentos a excluir
DOCS_EXCLUIR = {'35530070', '1070986776'}

INPUT_PATH  = r'c:\xampp\htdocs\Proyectos_React\bufala-bella\NominaSainco.xlsx'
OUTPUT_PATH = r'c:\xampp\htdocs\Proyectos_React\bufala-bella\NominaSainco_Procesada.xlsx'


# ─────────────────────────────────────────────────────────────────
# FUNCIÓN: días en meses de 30 días (base 30/360 colombiana) + 1
# ─────────────────────────────────────────────────────────────────
def dias_30_360_mas1(fecha_ini: date, fecha_fin: date) -> float:
    """Días entre dos fechas con meses de 30 días, sumando 1 día al resultado."""
    d1, m1, y1 = fecha_ini.day, fecha_ini.month, fecha_ini.year
    d2, m2, y2 = fecha_fin.day, fecha_fin.month, fecha_fin.year
    if d1 == 31:
        d1 = 30
    if d2 == 31:
        d2 = 30
    dias = (y2 - y1) * 360 + (m2 - m1) * 30 + (d2 - d1) + 1
    return max(dias, 0)


def redondear_mil_siguiente(valor: float) -> float:
    """Redondea al múltiplo de 1.000 inmediatamente superior."""
    return math.ceil(valor / 1000) * 1000


# ─────────────────────────────────────────────────────────────────
# 1. LEER ARCHIVO (detectar fila de encabezados automáticamente)
# ─────────────────────────────────────────────────────────────────
df_raw = pd.read_excel(INPUT_PATH, header=None)

header_row = None
for i, row in df_raw.iterrows():
    vals = [str(v).strip() for v in row.values]
    if any(v in ('QUINCENA', 'N. DOCUMENTO', 'NOMBRES') for v in vals):
        header_row = i
        break

if header_row is None:
    raise ValueError("No se encontró la fila de encabezados en el archivo.")

df_raw.columns = [str(v).strip() for v in df_raw.iloc[header_row].values]
df = df_raw.iloc[header_row + 1:].reset_index(drop=True)
df.dropna(how='all', inplace=True)

print("Columnas encontradas:", df.columns.tolist())
print(f"Total filas originales: {len(df)}")


# ─────────────────────────────────────────────────────────────────
# 2. EXCLUIR DOCUMENTOS
# ─────────────────────────────────────────────────────────────────
df['N. DOCUMENTO'] = df['N. DOCUMENTO'].astype(str).str.strip()
antes = len(df)
df = df[~df['N. DOCUMENTO'].isin(DOCS_EXCLUIR)].reset_index(drop=True)
print(f"Filas tras excluir documentos {DOCS_EXCLUIR}: {len(df)} (se eliminaron {antes - len(df)})\n")


# ─────────────────────────────────────────────────────────────────
# 3. NORMALIZAR FECHAS
# ─────────────────────────────────────────────────────────────────
df['Fecha Ingreso'] = pd.to_datetime(df['Fecha Ingreso'], errors='coerce')
df['Fecha Retiro']  = pd.to_datetime(df['Fecha Retiro'],  errors='coerce')
df['Pago']          = pd.to_numeric(df['Pago'], errors='coerce').fillna(0)


# ─────────────────────────────────────────────────────────────────
# 4. CONSOLIDAR POR N. DOCUMENTO (sumar Pago, mantener demás)
# ─────────────────────────────────────────────────────────────────
def agg_fecha_retiro(series):
    """Si algún valor es NaT → empleado activo → devuelve NaT."""
    if series.isna().any():
        return pd.NaT
    return series.max()

consolidado = df.groupby('N. DOCUMENTO', as_index=False).agg(
    QUINCENA      = ('QUINCENA',      'first'),
    NOMBRES       = ('NOMBRES',       'first'),
    Fecha_Ingreso = ('Fecha Ingreso', 'min'),
    Fecha_Retiro  = ('Fecha Retiro',  agg_fecha_retiro),
    Pago          = ('Pago',          'sum'),
)
print(f"Empleados únicos tras consolidar: {len(consolidado)}\n")


# ─────────────────────────────────────────────────────────────────
# 5. FECHAS EFECTIVAS Y DÍAS
# ─────────────────────────────────────────────────────────────────
def fecha_ini_efectiva(fi):
    if pd.isna(fi):
        return FECHA_MINIMA
    fi_d = fi.date() if hasattr(fi, 'date') else fi
    return max(fi_d, FECHA_MINIMA)

def fecha_fin_efectiva(fr):
    if pd.isna(fr):
        return FECHA_MAX_RETIRO
    fr_d = fr.date() if hasattr(fr, 'date') else fr
    return fr_d

consolidado['_FechaIni'] = consolidado['Fecha_Ingreso'].apply(fecha_ini_efectiva)
consolidado['_FechaFin'] = consolidado['Fecha_Retiro'].apply(fecha_fin_efectiva)
consolidado['_Dias']     = consolidado.apply(
    lambda r: dias_30_360_mas1(r['_FechaIni'], r['_FechaFin']), axis=1
)


# ─────────────────────────────────────────────────────────────────
# 6. COLUMNAS CALCULADAS
# ─────────────────────────────────────────────────────────────────
consolidado['Salario']      = SALARIO_DIARIO * consolidado['_Dias']
consolidado['AuxTransporte'] = AUX_DIARIO    * consolidado['_Dias']

# DctoSS → Salario * 8% redondeado al mil siguiente
consolidado['DctoSS'] = consolidado['Salario'].apply(
    lambda s: redondear_mil_siguiente(s * 0.08)
)

# Base (Salario + AuxTransporte)
consolidado['_Base'] = consolidado['Salario'] + consolidado['AuxTransporte']

# --- PrestacionesSociales bruta ---
prest_bruta = np.where(
    consolidado['_Base'] < consolidado['Pago'],
    consolidado['Pago'] - consolidado['_Base'],
    0.0
)

# Tope = Base * 21.83%
tope = consolidado['_Base'] * TOPE_PREST

# PrestacionesSociales: mínimo entre prest_bruta y el tope (redondeado)
prest_final = np.where(
    prest_bruta > tope,
    np.round(tope),
    np.round(prest_bruta)
)

# OtrosAuxilio: excedente cuando la prestación supera el tope
otros_auxilio = np.where(
    prest_bruta > tope,
    np.round(prest_bruta - np.round(tope)),
    0.0
)

consolidado['PrestacionesSociales'] = prest_final
consolidado['OtrosAuxilio']         = otros_auxilio

# DctoPermiso: si Base > Pago → Base - Pago, sino 0
consolidado['DctoPermiso'] = np.where(
    consolidado['_Base'] > consolidado['Pago'],
    np.round(consolidado['_Base'] - consolidado['Pago']),
    0.0
)


# ─────────────────────────────────────────────────────────────────
# 7. DATAFRAME FINAL
# ─────────────────────────────────────────────────────────────────
resultado = consolidado[[
    'N. DOCUMENTO', 'QUINCENA', 'NOMBRES',
    'Fecha_Ingreso', 'Fecha_Retiro',
    'Pago', 'Salario', 'AuxTransporte', 'DctoSS',
    'PrestacionesSociales', 'OtrosAuxilio', 'DctoPermiso',
]].copy()

resultado.rename(columns={
    'Fecha_Ingreso': 'Fecha Ingreso',
    'Fecha_Retiro':  'Fecha Retiro',
}, inplace=True)

# Redondear columnas monetarias a 2 decimales
cols_money = ['Pago', 'Salario', 'AuxTransporte', 'DctoSS',
              'PrestacionesSociales', 'OtrosAuxilio', 'DctoPermiso']
resultado[cols_money] = resultado[cols_money].round(2)


# ─────────────────────────────────────────────────────────────────
# 8. EXPORTAR A EXCEL CON FORMATO
# ─────────────────────────────────────────────────────────────────
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
    resultado.to_excel(writer, index=False, sheet_name='Nomina Procesada')
    ws = writer.sheets['Nomina Procesada']

    money_fmt = '#,##0.00'
    date_fmt  = 'DD/MM/YYYY'
    col_idx   = {cell.value: cell.column for cell in ws[1]}

    for col_name, fmt in [
        ('Pago',                 money_fmt),
        ('Salario',              money_fmt),
        ('AuxTransporte',        money_fmt),
        ('DctoSS',               money_fmt),
        ('PrestacionesSociales', money_fmt),
        ('OtrosAuxilio',         money_fmt),
        ('DctoPermiso',          money_fmt),
        ('Fecha Ingreso',        date_fmt),
        ('Fecha Retiro',         date_fmt),
    ]:
        if col_name in col_idx:
            for row in ws.iter_rows(min_row=2,
                                    min_col=col_idx[col_name],
                                    max_col=col_idx[col_name]):
                for cell in row:
                    cell.number_format = fmt

    # Encabezados
    header_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Auto-ancho
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)

print(f"✅ Archivo generado: {OUTPUT_PATH}")
print(f"   Total empleados procesados: {len(resultado)}\n")
print(resultado.to_string(index=False))
